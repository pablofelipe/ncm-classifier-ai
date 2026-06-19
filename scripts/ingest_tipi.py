# I/O adapter — parsing logic lives in src/core/domain/tipi_parsing.py.
# openpyxl chosen over pandas: precise row-number tracking (raw_row field),
# cell-level control for detecting partial subheadings, already in deps.
#
# Usage:
#   python scripts/ingest_tipi.py [chapter]   (default: 22)
#   Output: data/tipi/tipi_<chapter>_YYYYMMDD.json
import json
import re
import sys
from dataclasses import asdict
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.domain.tipi_parsing import RawRow, parse_tipi_rows  # noqa: E402

_BASE_DECREE_RE = re.compile(
    r"Decreto\s+n[º°.]\s*(\d+[\.\d]*),\s*de\s+\d+\s+de\s+\w+\s+de\s+(\d{4})"
)


def _fmt_rate(v: object) -> str:
    """Normalize openpyxl cell value to a clean rate string."""
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        return f"{round(v, 6):g}"  # 3.9000000000000004 -> "3.9"
    return str(v).strip()


def _extract_tipi_version(ws) -> str:  # type: ignore[no-untyped-def]
    """Build a human-readable version string from worksheet header rows."""
    row2_val = ws.cell(row=2, column=1).value or ""
    row3_val = ws.cell(row=3, column=1).value or ""

    base_match = _BASE_DECREE_RE.search(str(row2_val))
    base = (
        f"Decreto {base_match.group(1)}/{base_match.group(2)}"
        if base_match
        else "Decreto desconhecido"
    )

    update_lines = [ln.strip() for ln in str(row3_val).splitlines() if ln.strip()]
    last_update = update_lines[-1] if update_lines else ""

    return f"{base} (última atualização: {last_update})" if last_update else base


def _find_latest_xlsx(raw_dir: Path) -> Path:
    files = sorted(raw_dir.glob("tipi_*.xlsx"), reverse=True)
    if not files:
        raise FileNotFoundError(f"No tipi_*.xlsx found in {raw_dir}")
    return files[0]


def _load_raw_rows(xlsx_path: Path, chapter: str) -> tuple[list[RawRow], str]:
    """Open XLSX, walk rows, return (raw_rows_for_chapter, tipi_version)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    tipi_version = _extract_tipi_version(ws)
    rows: list[RawRow] = []
    in_chapter = False

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        ncm_cell = str(row[0]).strip() if row[0] is not None else ""

        if not ncm_cell or ncm_cell.upper().startswith("NCM"):
            continue

        if ncm_cell.startswith(chapter):
            in_chapter = True
        elif in_chapter:
            break  # past chapter boundary

        if not in_chapter:
            continue

        ex_raw = row[1]
        ex = str(ex_raw).strip() if ex_raw is not None else None

        rows.append(
            RawRow(
                ncm=ncm_cell,
                ex=ex if ex else None,
                description=str(row[2]).strip() if row[2] is not None else None,
                ipi_rate=_fmt_rate(row[3]),
                row_number=row_idx,
            )
        )

    wb.close()
    return rows, tipi_version


# v2 "beverage" corpus (ADR-0009): a curated multi-chapter index. Ch.22 whole;
# Ch.20 only heading 2009 (fruit/vegetable juices); Ch.21 only 2101 (coffee/tea
# extracts) + 2106.90.10 (beverage preparations). Prefixes are matched against
# the dotted NCM, so "2106.90.10" selects exactly that leaf.
BEVERAGE_FILTERS: dict[str, tuple[str, ...] | None] = {
    "22": None,  # whole chapter (34 NCMs)
    "20": ("2009",),  # juices (23 NCMs)
    "21": ("2101", "2106.90.10"),  # extracts + beverage preps (7 NCMs)
}


def ingest(raw_dir: Path, output_dir: Path, chapter: str = "22") -> Path:
    xlsx_path = _find_latest_xlsx(raw_dir)
    print(f"Lendo: {xlsx_path.name}")

    raw_rows, tipi_version = _load_raw_rows(xlsx_path, chapter)
    print(f"Linhas brutas Cap. {chapter}: {len(raw_rows)}")

    entries = parse_tipi_rows(raw_rows, chapter=chapter)
    print(f"Entradas NCM extraídas: {len(entries)}")

    date_tag = datetime.now(UTC).strftime("%Y%m%d")
    output_path = output_dir / f"tipi_{chapter}_{date_tag}.json"

    payload = {
        "tipi_version": tipi_version,
        "extracted_at": datetime.now(UTC).isoformat(),
        "chapter": chapter,
        "source": xlsx_path.name,
        "entries": [asdict(e) for e in entries],
    }

    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Escrito: {output_path.relative_to(ROOT)}  ({len(entries)} entradas)")

    version_file = ROOT / "eval" / "tipi_version.txt"
    version_file.write_text(xlsx_path.name + "\n", encoding="utf-8")

    return output_path


def ingest_beverage(raw_dir: Path, output_dir: Path) -> Path:
    """Build the curated multi-chapter beverage corpus (see BEVERAGE_FILTERS).

    Each chapter is parsed with the same chapter-agnostic parser used for the
    per-chapter files, then filtered by NCM prefix. Hierarchical enrichment
    (heading_description / subheading_description) is already attached by the
    parser, so post-parse filtering preserves it.
    """
    xlsx_path = _find_latest_xlsx(raw_dir)
    print(f"Lendo: {xlsx_path.name}")

    entries = []
    tipi_version = ""
    for chapter, prefixes in BEVERAGE_FILTERS.items():
        raw_rows, version = _load_raw_rows(xlsx_path, chapter)
        tipi_version = tipi_version or version
        parsed = parse_tipi_rows(raw_rows, chapter=chapter)
        if prefixes is not None:
            parsed = [e for e in parsed if any(e.ncm.startswith(p) for p in prefixes)]
        print(f"Cap. {chapter}: {len(parsed)} NCMs")
        entries.extend(parsed)

    # Fixed to the TIPI vigência date (not wall-clock) so the corpus filename is
    # reproducible and reflects the regulation in force, not when it was built.
    date_tag = "20260618"
    output_path = output_dir / f"tipi_beverage_{date_tag}.json"

    payload = {
        "tipi_version": tipi_version,
        "extracted_at": datetime.now(UTC).isoformat(),
        "chapter": "beverage",
        "chapters": list(BEVERAGE_FILTERS),
        "source": xlsx_path.name,
        "entries": [asdict(e) for e in entries],
    }

    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Escrito: {output_path.relative_to(ROOT)}  ({len(entries)} entradas)")
    return output_path


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "22"
    raw_dir = ROOT / "data" / "tipi" / "raw"
    output_dir = ROOT / "data" / "tipi"
    if target == "beverage":
        ingest_beverage(raw_dir=raw_dir, output_dir=output_dir)
    else:
        ingest(raw_dir=raw_dir, output_dir=output_dir, chapter=target)
